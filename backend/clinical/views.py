from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# Modelos y Serializers
from .models import Madre, Parto, RecienNacido, LogAudit, Alta, Perfil
from django.contrib.auth.models import User
from core.encryption import crypto_manager
from .serializers import *

# --- HELPER LOGS ---
def registrar_log(request, accion, modulo, descripcion):
    try:
        rol = 'Desconocido'
        if request.user.is_authenticated and hasattr(request.user, 'perfil'):
            rol = request.user.perfil.rol
        
        LogAudit.objects.create(
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
            rol=rol,
            accion=accion,
            modelo=modulo,
            detalles=descripcion,
            ip_address=request.META.get('REMOTE_ADDR')
        )
    except Exception as e:
        # Si falla el log, lo imprimimos en la consola negra para enterarnos
        print(f"❌ Error guardando LOG: {e}")

# --- VIEWSETS (Lógica API) ---
class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


class MadreViewSet(viewsets.ModelViewSet):
    queryset = Madre.objects.all().order_by('-created_at')
    serializer_class = MadreSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None
    # LOGICA DE GUARDADO
    def perform_create(self, serializer):
        ins = serializer.save()
        registrar_log(self.request, 'CREAR', 'Madre', f"ID: {ins.id}")

    def list(self, request, *args, **kwargs):
        rut_buscado = request.query_params.get('rut', None)
        # 1. SI ESTÁN BUSCANDO UN RUT ESPECÍFICO...
        if rut_buscado:
            input_limpio = rut_buscado.replace('.', '').replace('-', '').strip().upper()
            # Buscamos una por una desencriptando
            for madre in Madre.objects.all():
                try:
                    rut_real = crypto_manager.decrypt(madre.rut)
                    if rut_real:
                        rut_bd_limpio = rut_real.replace('.', '').replace('-', '').strip().upper()
                        
                        if input_limpio == rut_bd_limpio:
                            serializer = self.get_serializer(madre)
                            return Response([serializer.data])
                except:
                    continue
            return Response([]) 
        
        return super().list(request, *args, **kwargs)

class PartoViewSet(viewsets.ModelViewSet):
    queryset = Parto.objects.all().order_by('-fecha')
    serializer_class = PartoSerializer
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        ins = serializer.save()
        registrar_log(self.request, 'CREAR', 'Parto', f"ID: {ins.id}")

class RecienNacidoViewSet(viewsets.ModelViewSet):
    queryset = RecienNacido.objects.all()
    serializer_class = RecienNacidoSerializer
    permission_classes = [IsAuthenticated]

class AltaViewSet(viewsets.ModelViewSet):
    queryset = Alta.objects.all().order_by('-fecha_solicitud')
    serializer_class = AltaSerializer
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        serializer.save() 
        registrar_log(self.request, 'CREAR', 'Alta', "Solicitud")
    @action(detail=True, methods=['patch'])
    def gestionar(self, request, pk=None):
        alta = self.get_object(); alta.estado = request.data.get('estado'); alta.autorizado_por = request.user; alta.save()
        registrar_log(request, 'ACTUALIZAR', 'Alta', f"Estado: {alta.estado}")
        return Response({'status': 'ok'})

class LogAuditViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = LogAudit.objects.all().order_by('-fecha')
    serializer_class = LogAuditSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('id')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    @action(detail=True, methods=['post'])
    def toggle_estado(self, request, pk=None):
        usuario = self.get_object()
        # Pequeña protección: Evitar que el usuario se bloquee a sí mismo por error
        if usuario.id == request.user.id:
            return Response({'error': 'No puedes bloquear tu propio usuario'}, status=400)

        usuario.is_active = not usuario.is_active
        usuario.save()
        
        estado_texto = "Activo" if usuario.is_active else "Bloqueado"
        
        try:
            registrar_log(request, 'ACTUALIZAR', 'Usuario', f"Cambió estado a {estado_texto} - UserID: {usuario.id}")
        except:
            pass # Si falla el log, que no se caiga la app
            
        return Response({
            'status': 'success', 
            'is_active': usuario.is_active
        })

    # --- REPORTES ---
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_excel_completo(request):
    registrar_log(request, 'EXPORTAR', 'Excel', "Listado Partos Estilizado")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Gestion_Partos.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro Clínico"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    centered_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- 2. ENCABEZADOS ---
    headers = ['ID', 'Fecha Ingreso', 'Nombre Paciente', 'RUT', 'Tipo Parto', 'Sexo RN', 'Peso (gr)']
    ws.append(headers)
    
    # Aplicar estilo a la fila de encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = thin_border

    # --- 3. DATOS (Con Desencriptación) ---
    for p in Parto.objects.all().order_by('-fecha'):
        # Desencriptar
        rut_visible = "Sin Dato"
        try:
            rut_visible = crypto_manager.decrypt(p.madre.rut) if p.madre.rut else ""
        except: rut_visible = "Error Cifrado"

        nombre_visible = "Sin Nombre"
        try:
            if p.madre.nombre_completo:
                decrypted = crypto_manager.decrypt(p.madre.nombre_completo)
                nombre_visible = decrypted if decrypted != "ERROR_DECRYPT" else p.madre.nombre_completo
            else:
                nombre_visible = p.madre.nombre_completo
        except: nombre_visible = p.madre.nombre_completo

        rn = p.recien_nacidos.first()
        sexo = rn.sexo if rn else "-"
        peso = rn.peso_gramos if rn else 0

        row = [
            p.id,
            str(p.fecha)[:16],
            nombre_visible,
            rut_visible,
            p.tipo_parto,
            sexo,
            peso
        ]
        ws.append(row)

    column_widths = [5, 20, 35, 15, 15, 10, 10]
    
    for i, column_width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.column == 3: 
                cell.alignment = left_alignment
            else:
                cell.alignment = centered_alignment

    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_pdf_rem(request):
    # 1. Estadísticas Generales
    total = Parto.objects.count()
    cesareas = Parto.objects.filter(tipo_parto__icontains='CESAREA').count()
    normales = total - cesareas
    
    # 2. PROCESAMIENTO DE DATOS (Limpieza forense)
    partos_db = Parto.objects.all().order_by('-fecha')[:50]
    lista_partos_limpia = []

    for p in partos_db:
        # A. Desencriptamos RUT
        rut_visible = "Error"
        try:
            if p.madre.rut:
                rut_visible = crypto_manager.decrypt(p.madre.rut)
        except:
            rut_visible = "RUT Ilegible"

        # B. Desencriptamos NOMBRE (¡ESTO FALTABA!)
        nombre_visible = "Desconocido"
        try:
            if p.madre.nombre_completo:
                decrypted_name = crypto_manager.decrypt(p.madre.nombre_completo)
                # Si desencripta bien, usamos ese. Si no (porque era texto plano antiguo), usamos el original.
                if decrypted_name and decrypted_name != "ERROR_DECRYPT":
                    nombre_visible = decrypted_name
                else:
                    nombre_visible = p.madre.nombre_completo
        except:
            # Si falla, mostramos el original (aunque esté feo, mejor que nada)
            nombre_visible = p.madre.nombre_completo

        # C. Datos del Bebé
        rn = p.recien_nacidos.first()
        sexo_rn = rn.sexo if rn else "S/I"
        peso_rn = f"{rn.peso_gramos} gr" if rn else "0 gr"

        # D. Armamos la fila limpia
        lista_partos_limpia.append({
            'fecha': p.fecha,
            'tipo': p.tipo_parto,
            'profesional': "Matrona Turno",
            'rut_madre': rut_visible,       # <--- Limpio
            'nombre_madre': nombre_visible, # <--- ¡AHORA SÍ LIMPIO!
            'sexo': sexo_rn,
            'peso': peso_rn
        })
    
    # 3. Contexto
    context = {
        'lista_partos': lista_partos_limpia,
        'total_partos': total,
        'total_cesareas': cesareas,
        'total_normales': normales,
        'fecha_generacion': timezone.now(),
        'generado_por': request.user.username
    }
    
    # 4. Generar PDF
    template_path = 'reporte_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Informe_REM_Gestion.pdf"'
    
    try:
        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error creando PDF', status=500)
    except Exception as e:
        return HttpResponse(f'Error en plantilla: {e}', status=500)
        
    registrar_log(request, 'EXPORTAR', 'PDF', "Reporte REM (Limpio)")
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def alta_medica_pdf(request, pk):
    alta = get_object_or_404(Alta, pk=pk)
    
    # 1. Preparar datos desencriptados por defecto
    rut_visible = "RUT No Disponible"
    paciente_nombre = "Nombre No Disponible"
    
    # 2. Intentar desencriptar RUT
    if alta.parto.madre.rut:
        try:
            rut_visible = crypto_manager.decrypt(alta.parto.madre.rut)
        except:
            rut_visible = "Error Desencriptado"

    # 3. Intentar desencriptar Nombre (si tu modelo lo guarda encriptado)
    if alta.parto.madre.nombre_completo:
        try:
            # Intentamos desencriptar
            decrypted_name = crypto_manager.decrypt(alta.parto.madre.nombre_completo)
            if decrypted_name and decrypted_name != "ERROR_DECRYPT":
                paciente_nombre = decrypted_name
            else:
                # Si no era encriptado, usamos el original
                paciente_nombre = alta.parto.madre.nombre_completo
        except:
            paciente_nombre = alta.parto.madre.nombre_completo
    # 4. Contexto que recibe el HTML
    context = {
        'tipo_alta': alta.tipo,
        'paciente_nombre': paciente_nombre,  # <--- VARIABLE LIMPIA
        'rut_visible': rut_visible,          # <--- VARIABLE LIMPIA
        'fecha_ingreso': alta.parto.fecha,
        'responsable': alta.autorizado_por.username if alta.autorizado_por else "Sin Firma",
        'fecha_generacion': timezone.now(),
        'ip': request.META.get('REMOTE_ADDR')
    }
    # 5. Generar PDF
    template_path = 'alta_medica.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Alta_{pk}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error generando PDF')
        
    registrar_log(request, 'EXPORTAR', 'PDF', f"Alta {pk} (HTML)")
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_auditoria_pdf(request):
    # 1. Obtenemos los últimos 100 registros de actividad
    logs = LogAudit.objects.all().order_by('-fecha')[:100]
    # 2. Contexto para el HTML
    context = {
        'logs': logs,
        'fecha_generacion': timezone.now(),
        'solicitante': request.user.username,
        'total_registros': logs.count()
    }
    # 3. Generar PDF usando plantilla
    template_path = 'auditoria_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Auditoria_Forense_Sistema.pdf"'
    
    try:
        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error creando PDF', status=500)
    except Exception as e:
        return HttpResponse(f'Error buscando plantilla: {e}', status=500)
        
    registrar_log(request, 'EXPORTAR', 'PDF', "Auditoria de Seguridad")
    return response